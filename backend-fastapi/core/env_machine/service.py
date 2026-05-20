#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Author: 臧成龙
@Contact: 939589097@qq.com
@Time: 2025-03-25
@File: service.py
@Desc: 执行机基础服务层 - CRUD 操作
"""
import json
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, Alignment
from pydantic import BaseModel
from sqlalchemy import select, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.base_service import BaseService
from app.config import get_settings
from core.env_machine.model import EnvMachine

settings = get_settings()


class EnvMachineCreateSchema(BaseModel):
    """
    创建 Schema 占位类

    由于执行机的创建主要通过注册接口（EnvRegisterRequest）完成，
    这里使用占位类以满足 BaseService 的泛型要求。
    """
    pass


class EnvMachineUpdateSchema(BaseModel):
    """
    更新 Schema 占位类

    执行机的更新主要通过特定的业务接口完成，
    这里使用占位类以满足 BaseService 的泛型要求。
    """
    pass


class EnvMachineService(BaseService[EnvMachine, EnvMachineCreateSchema, EnvMachineUpdateSchema]):
    """
    执行机服务层

    继承 BaseService，自动获得以下基础功能：
    - create(): 创建记录
    - get_by_id(): 根据 ID 获取单条记录
    - get_list(): 获取分页列表
    - update(): 更新记录
    - delete(): 删除记录（支持软删除）
    - batch_delete(): 批量删除
    - export_to_excel(): 导出 Excel
    - import_from_excel(): 导入 Excel
    - check_unique(): 检查字段唯一性
    - get_by_field(): 根据字段获取单条记录
    - exists(): 检查记录是否存在
    """

    model = EnvMachine

    # Excel 导入导出配置
    excel_columns = {
        "namespace": "机器分类",
        "ip": "机器IP",
        "port": "端口",
        "asset_number": "资产编号",
        "device_type": "机器类型",
        "device_sn": "设备SN",
        "mark": "标签",
        "available": "是否启用",
        "status": "状态",
        "version": "版本",
        "note": "备注",
    }
    excel_sheet_name = "执行机列表"

    # 虚拟设备 Excel 导入导出配置
    VIRTUAL_EXCEL_COLUMNS = {
        "namespace": "机器分类",
        "device_type": "机器类型",
        "asset_number": "资产编号",
        "ip": "虚拟标识",
        "mark": "标签",
        "extra_message": "扩展信息(JSON)",
        "note": "备注",
    }

    @classmethod
    def _export_converter(cls, item: EnvMachine) -> Dict[str, Any]:
        """导出数据转换器"""
        return {
            "namespace": item.namespace or "",
            "ip": item.ip or "",
            "port": item.port or "",
            "asset_number": item.asset_number or "",
            "device_type": item.device_type or "",
            "device_sn": item.device_sn or "",
            "mark": item.mark or "",
            "available": "是" if item.available else "否",
            "status": item.get_status_display(),
            "version": item.version or "",
            "note": item.note or "",
        }

    @classmethod
    def _import_processor(cls, row: Dict[str, Any]) -> Optional[EnvMachine]:
        """导入数据处理器"""
        namespace = row.get("namespace")
        ip = row.get("ip")
        port = row.get("port")
        device_type = row.get("device_type")

        # 必填字段校验
        if not all([namespace, ip, port, device_type]):
            return None

        available_str = row.get("available", "是")
        available = available_str in ("是", "true", "True", "1", True)

        status_str = row.get("status", "在线")
        status_map = {"在线": "online", "使用中": "using", "离线": "offline"}
        status = status_map.get(status_str, "online")

        return EnvMachine(
            namespace=str(namespace),
            ip=str(ip),
            port=str(port),
            asset_number=str(row.get("asset_number")) if row.get("asset_number") else None,
            device_type=str(device_type),
            device_sn=str(row.get("device_sn")) if row.get("device_sn") else None,
            mark=str(row.get("mark")) if row.get("mark") else None,
            available=available,
            status=status,
            version=str(row.get("version")) if row.get("version") else None,
            note=str(row.get("note")) if row.get("note") else None,
        )

    @classmethod
    def _virtual_import_processor(cls, row: Dict[str, Any]) -> Optional[EnvMachine]:
        """虚拟设备导入处理器"""
        namespace = row.get("namespace")
        device_type = row.get("device_type")
        asset_number = row.get("asset_number")
        ip = row.get("ip")
        mark = row.get("mark")
        extra_message_str = row.get("extra_message")

        if not all([namespace, device_type, asset_number, ip, mark, extra_message_str]):
            return None

        try:
            extra_message = json.loads(extra_message_str)
            if not isinstance(extra_message, dict):
                return None
        except json.JSONDecodeError:
            return None

        return EnvMachine(
            namespace=str(namespace),
            device_type=str(device_type),
            asset_number=str(asset_number),
            ip=str(ip),
            port=None,
            device_sn=None,
            mark=str(mark),
            available=False,
            status="online",
            is_virtual=True,
            extra_message=extra_message,
            note=str(row.get("note")) if row.get("note") else None,
        )

    @classmethod
    async def export_to_excel(
        cls,
        db: AsyncSession,
        data_converter: Any = None
    ) -> BytesIO:
        """导出到 Excel"""
        return await super().export_to_excel(db, cls._export_converter)

    @classmethod
    async def import_from_excel(
        cls,
        db: AsyncSession,
        file_content: bytes,
        row_processor: Any = None
    ) -> Tuple[int, int]:
        """从 Excel 导入"""
        return await super().import_from_excel(db, file_content, cls._import_processor)

    @classmethod
    async def import_virtual_from_excel(
        cls,
        db: AsyncSession,
        file_content: bytes,
    ) -> Tuple[int, List[Dict[str, Any]]]:
        """从 Excel 导入虚拟设备"""
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(file_content))
        ws = wb.active

        success_count = 0
        failed_items = []

        headers = [cell.value for cell in ws[1]]
        column_map = {}
        for key, cn_name in cls.VIRTUAL_EXCEL_COLUMNS.items():
            for idx, header in enumerate(headers):
                if header == cn_name:
                    column_map[key] = idx
                    break

        required_keys = ["namespace", "device_type", "asset_number", "ip", "mark", "extra_message"]
        missing_keys = [k for k in required_keys if k not in column_map]
        if missing_keys:
            return 0, [{"row": 0, "reason": f"缺少必填列: {cls.VIRTUAL_EXCEL_COLUMNS[k]}" for k in missing_keys}]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            row_dict = {}
            for key, idx in column_map.items():
                row_dict[key] = row[idx] if idx < len(row) else None

            machine = cls._virtual_import_processor(row_dict)
            if machine:
                db.add(machine)
                success_count += 1
            else:
                failed_items.append({
                    "row": row_idx,
                    "reason": "数据验证失败（必填字段缺失或 JSON 格式错误）"
                })

        if success_count > 0:
            await db.commit()

        return success_count, failed_items

    @classmethod
    async def get_by_namespace(
        cls,
        db: AsyncSession,
        namespace: str,
        page: int = 1,
        page_size: int = 100
    ) -> Tuple[List[EnvMachine], int]:
        """
        根据 namespace 获取机器列表

        :param db: 数据库会话
        :param namespace: 机器分类
        :param page: 页码
        :param page_size: 每页数量
        :return: (机器列表, 总数)
        """
        filters = [EnvMachine.namespace == namespace]
        return await cls.get_list(db, page=page, page_size=page_size, filters=filters)

    @classmethod
    async def get_list_with_filters(
        cls,
        db: AsyncSession,
        namespace: Optional[str] = None,  # 改为 Optional
        device_type: Optional[str] = None,
        ip: Optional[str] = None,
        asset_number: Optional[str] = None,
        mark: Optional[str] = None,
        available: Optional[bool] = None,
        note: Optional[str] = None,
        page: int = 1,
        page_size: int = 20
    ) -> Tuple[List[EnvMachine], int]:
        """
        多条件查询执行机列表

        :param db: 数据库会话
        :param namespace: 机器分类（可选，None表示查询全部）
        :param device_type: 机器类型
        :param ip: IP地址（模糊查询）
        :param asset_number: 资产编号（模糊查询）
        :param mark: 标签（模糊查询）
        :param available: 是否启用
        :param note: 备注（模糊查询）
        :param page: 页码
        :param page_size: 每页数量
        :return: (机器列表, 总数)
        """
        # 定义有效的命名空间列表（排除手工使用，从配置动态获取）
        VALID_NAMESPACES = list(settings.namespace_map.keys())

        filters = [EnvMachine.is_deleted == False]

        if namespace:
            filters.append(EnvMachine.namespace == namespace)
        else:
            # namespace 为 None 时，查询所有4个命名空间
            filters.append(EnvMachine.namespace.in_(VALID_NAMESPACES))

        if device_type:
            filters.append(EnvMachine.device_type == device_type)

        if ip:
            escaped_ip = ip.replace("%", r"\%").replace("_", r"\_")
            filters.append(EnvMachine.ip.ilike(f"%{escaped_ip}%"))

        if asset_number:
            escaped_asset_number = asset_number.replace("%", r"\%").replace("_", r"\_")
            filters.append(EnvMachine.asset_number.ilike(f"%{escaped_asset_number}%"))

        if mark:
            escaped_mark = mark.replace("%", r"\%").replace("_", r"\_")
            filters.append(EnvMachine.mark.ilike(f"%{escaped_mark}%"))

        if available is not None:
            filters.append(EnvMachine.available == available)

        if note:
            escaped_note = note.replace("%", r"\%").replace("_", r"\_")
            filters.append(EnvMachine.note.ilike(f"%{escaped_note}%"))

        return await cls.get_list(db, page=page, page_size=page_size, filters=filters)

    @classmethod
    async def get_online_machines(
        cls,
        db: AsyncSession,
        namespace: Optional[str] = None,
        device_type: Optional[str] = None,
        page: int = 1,
        page_size: int = 100
    ) -> Tuple[List[EnvMachine], int]:
        """
        获取在线机器列表

        :param db: 数据库会话
        :param namespace: 可选，按分类筛选
        :param device_type: 可选，按设备类型筛选
        :param page: 页码
        :param page_size: 每页数量
        :return: (机器列表, 总数)
        """
        filters = [
            EnvMachine.status == "online",
            EnvMachine.available == True,  # noqa: E712
        ]

        if namespace:
            filters.append(EnvMachine.namespace == namespace)

        if device_type:
            filters.append(EnvMachine.device_type == device_type)

        return await cls.get_list(db, page=page, page_size=page_size, filters=filters)

    @classmethod
    async def get_by_ip(
        cls,
        db: AsyncSession,
        ip: str,
        namespace: Optional[str] = None
    ) -> Optional[EnvMachine]:
        """
        根据 IP 获取机器

        :param db: 数据库会话
        :param ip: 机器 IP
        :param namespace: 可选，同时按分类筛选
        :return: 机器记录或 None
        """
        query = select(EnvMachine).where(
            EnvMachine.ip == ip,
            EnvMachine.is_deleted == False  # noqa: E712
        )

        if namespace:
            query = query.where(EnvMachine.namespace == namespace)

        result = await db.execute(query)
        return result.scalar_one_or_none()

    @classmethod
    async def get_by_device_sn(
        cls,
        db: AsyncSession,
        device_sn: str,
        namespace: Optional[str] = None
    ) -> Optional[EnvMachine]:
        """
        根据设备 SN 获取机器

        :param db: 数据库会话
        :param device_sn: 设备序列号
        :param namespace: 可选，同时按分类筛选
        :return: 机器记录或 None
        """
        query = select(EnvMachine).where(
            EnvMachine.device_sn == device_sn,
            EnvMachine.is_deleted == False  # noqa: E712
        )

        if namespace:
            query = query.where(EnvMachine.namespace == namespace)

        result = await db.execute(query)
        return result.scalar_one_or_none()

    @classmethod
    async def get_by_namespace_and_device(
        cls,
        db: AsyncSession,
        namespace: str,
        ip: str,
        device_type: str,
        device_sn: Optional[str] = None
    ) -> Optional[EnvMachine]:
        """
        根据命名空间、IP、设备类型和设备SN获取机器（唯一标识查询）

        :param db: 数据库会话
        :param namespace: 机器分类
        :param ip: 机器 IP
        :param device_type: 设备类型
        :param device_sn: 设备 SN（移动端必填）
        :return: 机器记录或 None
        """
        query = select(EnvMachine).where(
            EnvMachine.namespace == namespace,
            EnvMachine.ip == ip,
            EnvMachine.device_type == device_type,
            EnvMachine.is_deleted == False  # noqa: E712
        )

        if device_sn:
            query = query.where(EnvMachine.device_sn == device_sn)

        result = await db.execute(query)
        return result.scalar_one_or_none()

    @classmethod
    async def search(
        cls,
        db: AsyncSession,
        keyword: str,
        namespace: Optional[str] = None,
        page: int = 1,
        page_size: int = 20
    ) -> Tuple[List[EnvMachine], int]:
        """
        搜索执行机（模糊匹配 IP、标签或备注）

        :param db: 数据库会话
        :param keyword: 搜索关键词
        :param namespace: 可选，按分类筛选
        :param page: 页码
        :param page_size: 每页数量
        :return: (机器列表, 总数)
        """
        # 转义 SQL LIKE 特殊字符，防止用户输入的 % 和 _ 被当作通配符
        escaped_keyword = keyword.replace("%", r"\%").replace("_", r"\_")
        pattern = f"%{escaped_keyword}%"

        filters = [
            or_(
                EnvMachine.ip.ilike(pattern),
                EnvMachine.mark.ilike(pattern),
                EnvMachine.note.ilike(pattern),
            )
        ]

        if namespace:
            filters.append(EnvMachine.namespace == namespace)

        return await cls.get_list(db, page=page, page_size=page_size, filters=filters)

    @classmethod
    async def update_status(
        cls,
        db: AsyncSession,
        machine_id: str,
        status: str,
        auto_commit: bool = True
    ) -> Optional[EnvMachine]:
        """
        更新机器状态

        :param db: 数据库会话
        :param machine_id: 机器 ID
        :param status: 新状态（online/using/offline）
        :param auto_commit: 是否自动提交
        :return: 更新后的机器或 None
        """
        valid_statuses = ["online", "using", "offline"]
        if status not in valid_statuses:
            raise ValueError(f"无效的状态值: {status}，有效值为: {valid_statuses}")

        machine = await cls.get_by_id(db, machine_id)
        if not machine:
            return None

        machine.status = status

        if auto_commit:
            await db.commit()
            await db.refresh(machine)
        else:
            await db.flush()
            await db.refresh(machine)

        return machine

    @classmethod
    async def update_available(
        cls,
        db: AsyncSession,
        machine_id: str,
        available: bool,
        auto_commit: bool = True
    ) -> Optional[EnvMachine]:
        """
        更新机器启用状态

        :param db: 数据库会话
        :param machine_id: 机器 ID
        :param available: 是否启用
        :param auto_commit: 是否自动提交
        :return: 更新后的机器或 None
        """
        machine = await cls.get_by_id(db, machine_id)
        if not machine:
            return None

        machine.available = available

        if auto_commit:
            await db.commit()
            await db.refresh(machine)
        else:
            await db.flush()
            await db.refresh(machine)

        return machine

    @classmethod
    async def batch_update_status(
        cls,
        db: AsyncSession,
        ids: List[str],
        status: str
    ) -> Tuple[int, List[str]]:
        """
        批量更新机器状态

        :param db: 数据库会话
        :param ids: 机器 ID 列表
        :param status: 新状态
        :return: (成功数量, 失败的 ID 列表)
        """
        valid_statuses = ["online", "using", "offline"]
        if status not in valid_statuses:
            raise ValueError(f"无效的状态值: {status}，有效值为: {valid_statuses}")

        success_count = 0
        failed_ids = []

        for machine_id in ids:
            machine = await cls.get_by_id(db, machine_id)
            if machine:
                machine.status = status
                success_count += 1
            else:
                failed_ids.append(machine_id)

        if success_count > 0:
            await db.commit()

        return success_count, failed_ids

    @classmethod
    async def batch_update_available(
        cls,
        db: AsyncSession,
        ids: List[str],
        available: bool
    ) -> Tuple[int, List[str]]:
        """
        批量更新机器启用状态

        :param db: 数据库会话
        :param ids: 机器 ID 列表
        :param available: 是否启用
        :return: (成功数量, 失败的 ID 列表)
        """
        success_count = 0
        failed_ids = []

        for machine_id in ids:
            machine = await cls.get_by_id(db, machine_id)
            if machine:
                machine.available = available
                success_count += 1
            else:
                failed_ids.append(machine_id)

        if success_count > 0:
            await db.commit()

        return success_count, failed_ids

    @classmethod
    async def get_stats(cls, db: AsyncSession, namespace: Optional[str] = None) -> Dict[str, Any]:
        """
        获取执行机统计信息

        :param db: 数据库会话
        :param namespace: 可选，按分类筛选
        :return: 统计信息字典
        """
        from sqlalchemy import func

        base_filter = [EnvMachine.is_deleted == False]  # noqa: E712
        if namespace:
            base_filter.append(EnvMachine.namespace == namespace)

        # 总数
        total_query = select(func.count(EnvMachine.id)).where(*base_filter)
        total_result = await db.execute(total_query)
        total_count = total_result.scalar() or 0

        # 按状态统计
        status_stats = {}
        for status in ["online", "using", "offline"]:
            status_query = select(func.count(EnvMachine.id)).where(
                *base_filter,
                EnvMachine.status == status
            )
            status_result = await db.execute(status_query)
            status_stats[status] = status_result.scalar() or 0

        # 按设备类型统计
        type_stats = {}
        for device_type in ["windows", "mac", "android", "ios"]:
            type_query = select(func.count(EnvMachine.id)).where(
                *base_filter,
                EnvMachine.device_type == device_type
            )
            type_result = await db.execute(type_query)
            type_stats[device_type] = type_result.scalar() or 0

        # 启用/禁用统计
        available_query = select(func.count(EnvMachine.id)).where(
            *base_filter,
            EnvMachine.available == True  # noqa: E712
        )
        available_result = await db.execute(available_query)
        available_count = available_result.scalar() or 0

        return {
            "total_count": total_count,
            "available_count": available_count,
            "unavailable_count": total_count - available_count,
            "status_stats": status_stats,
            "type_stats": type_stats,
        }

    @classmethod
    async def get_namespaces(cls, db: AsyncSession) -> List[str]:
        """
        获取所有机器分类（去重，排除包含 manual 的分类）

        :param db: 数据库会话
        :return: 分类列表
        """
        from sqlalchemy import distinct

        result = await db.execute(
            select(distinct(EnvMachine.namespace)).where(
                EnvMachine.is_deleted == False,  # noqa: E712
                EnvMachine.namespace.notlike('%manual%')  # 排除包含 manual 的 namespace
            ).order_by(EnvMachine.namespace)
        )
        return [row[0] for row in result.all()]

    @classmethod
    async def get_device_types(cls, db: AsyncSession, namespace: Optional[str] = None) -> List[str]:
        """
        获取所有设备类型（去重）

        :param db: 数据库会话
        :param namespace: 可选，按分类筛选
        :return: 设备类型列表
        """
        from sqlalchemy import distinct

        query = select(distinct(EnvMachine.device_type)).where(
            EnvMachine.is_deleted == False  # noqa: E712
        )

        if namespace:
            query = query.where(EnvMachine.namespace == namespace)

        query = query.order_by(EnvMachine.device_type)

        result = await db.execute(query)
        return [row[0] for row in result.all()]

    @classmethod
    async def generate_virtual_import_template(cls) -> BytesIO:
        """生成虚拟设备导入 Excel 模板"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "虚拟设备导入模板"

        headers = list(cls.VIRTUAL_EXCEL_COLUMNS.values())
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        ws.append([
            "meeting_virtual",
            "windows",
            "PERF-001",
            "perf001",
            "windows_perf",
            '{"windows_perf": {"username": "test", "password": "xxx"}}',
            "性能测试账号",
        ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer