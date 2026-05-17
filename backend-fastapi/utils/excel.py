#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Author: 臧成龙
@Contact: 939589097@qq.com
@Time: 2025-12-31
@File: excel.py
@Desc: Excel处理工具类 -
"""
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, Type, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pydantic import BaseModel

# 导出临时目录
TEMP_EXPORTS_DIR = Path("temp_exports")


class ExcelHandler:
    """Excel处理工具类"""

    # 表头样式
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    # 边框样式
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 最佳/最差值高亮样式
    BEST_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    WORST_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    @staticmethod
    def sanitize_sheet_name(version_name: str, suffix: str) -> str:
        """截断版本名，确保总长度 ≤ 31 字符（Excel 限制）"""
        max_version_len = 31 - len(suffix) - 1
        if len(version_name) > max_version_len:
            version_name = version_name[:max_version_len]
        return f"{version_name}-{suffix}" if suffix else version_name

    @classmethod
    def _write_table(cls, ws, start_row: int, title: str, data: List[Dict], columns: List[str]):
        """写入一个表格（带标题行和数据行）"""
        # 标题行
        title_row = start_row
        ws.cell(row=title_row, column=1, value=title).font = Font(bold=True, size=12)

        # 表头行
        header_row = start_row + 1
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.HEADER_ALIGNMENT
            cell.border = cls.THIN_BORDER

        # 数据行
        for row_idx, row_data in enumerate(data, header_row + 1):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(vertical="center")

    @classmethod
    def _highlight_extreme_values(cls, ws, metric_columns: List[str], start_row: int, end_row: int):
        """高亮最佳/最差值"""
        for col_idx, col_name in enumerate(metric_columns, start=4):
            values = [(row, ws.cell(row=row, column=col_idx).value) for row in range(start_row, end_row + 1)]
            valid_values = [(row, v) for row, v in values if v is not None]

            if not valid_values:
                continue

            # 判断规则：CPU/GPU/内存越小越好
            is_lower_better = any(kw in col_name for kw in ['CPU', 'GPU', '内存', '提交'])

            if is_lower_better:
                best_row = min(valid_values, key=lambda x: x[1])[0]
                worst_row = max(valid_values, key=lambda x: x[1])[0]
            else:
                best_row = max(valid_values, key=lambda x: x[1])[0]
                worst_row = min(valid_values, key=lambda x: x[1])[0]

            ws.cell(row=best_row, column=col_idx).fill = cls.BEST_FILL
            ws.cell(row=worst_row, column=col_idx).fill = cls.WORST_FILL

    @classmethod
    def write_summary_sheet(cls, ws, summary_data: 'SummaryData'):
        """写入摘要页"""
        # 基本信息（第1-2行）
        cls._write_table(ws, 1, "基本信息", summary_data.basic_info,
                         ["版本名称", "采集开始时间", "采集结束时间", "采集时长(秒)"])

        # 冲高区间（第5-6行）
        if summary_data.peak_range:
            peak_columns = ["版本名称", "冲高区间开始时间", "冲高区间结束时间"] + summary_data.metric_columns
            cls._write_table(ws, 5, "冲高区间峰值", summary_data.peak_range, peak_columns)
            cls._highlight_extreme_values(ws, summary_data.metric_columns, 6, 6 + len(summary_data.peak_range) - 1)

        # 稳态区间（第9-10行）
        if summary_data.steady_range:
            steady_columns = ["版本名称", "稳态区间开始时间", "稳态区间结束时间"] + summary_data.metric_columns
            cls._write_table(ws, 9, "稳态区间平均值", summary_data.steady_range, steady_columns)
            cls._highlight_extreme_values(ws, summary_data.metric_columns, 10, 10 + len(summary_data.steady_range) - 1)

    @classmethod
    def write_detail_sheet(cls, ws, detail_data: 'DetailData'):
        """写入详细数据页"""
        # 表头
        for col_idx, col_name in enumerate(detail_data.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.border = cls.THIN_BORDER

        # 数据行
        for row_idx, row_data in enumerate(detail_data.data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = cls.THIN_BORDER

    @classmethod
    def create_compare_excel(
        cls,
        summary_data: 'SummaryData',
        detail_data: Dict[str, 'DetailData'],
        metric: str,
        hwinfo_key: Optional[str] = None
    ) -> str:
        """创建版本对比 Excel 文件，返回文件路径"""
        # 确保临时目录存在
        TEMP_EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Sheet 1: 摘要页
        ws_summary = wb.active
        ws_summary.title = "数据摘要"
        cls.write_summary_sheet(ws_summary, summary_data)

        # Sheet 2~N: 详细数据页
        for sheet_name, data in detail_data.items():
            safe_name = cls.sanitize_sheet_name(sheet_name, "")
            ws = wb.create_sheet(title=safe_name)
            cls.write_detail_sheet(ws, data)

        # 保存到临时目录
        file_name = f"版本对比报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = TEMP_EXPORTS_DIR / file_name
        wb.save(file_path)

        return str(file_path)
    
    @classmethod
    def export_to_excel(
        cls,
        data: List[Dict[str, Any]],
        columns: Dict[str, str],
        sheet_name: str = "Sheet1"
    ) -> BytesIO:
        """
        导出数据到Excel
        :param data: 数据列表，每个元素是一个字典
        :param columns: 列映射，格式为 {字段名: 显示名}
        :param sheet_name: 工作表名称
        :return: Excel文件的BytesIO对象
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 写入表头
        headers = list(columns.values())
        field_names = list(columns.keys())
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.HEADER_ALIGNMENT
            cell.border = cls.THIN_BORDER
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field in enumerate(field_names, 1):
                value = row_data.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(vertical="center")
        
        # 自动调整列宽
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @classmethod
    def import_from_excel(
        cls,
        file_content: bytes,
        columns: Dict[str, str],
        schema: Optional[Type[BaseModel]] = None
    ) -> List[Dict[str, Any]]:
        """
        从Excel导入数据
        :param file_content: Excel文件内容
        :param columns: 列映射，格式为 {字段名: 显示名}
        :param schema: 可选的Pydantic Schema用于数据验证
        :return: 数据列表
        """
        wb = load_workbook(filename=BytesIO(file_content), read_only=True)
        ws = wb.active
        
        # 读取表头，建立显示名到字段名的映射
        header_to_field = {v: k for k, v in columns.items()}
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        
        # 第一行是表头
        headers = rows[0]
        field_indices = {}
        for idx, header in enumerate(headers):
            if header in header_to_field:
                field_indices[idx] = header_to_field[header]
        
        # 读取数据行
        result = []
        for row in rows[1:]:
            if not any(row):  # 跳过空行
                continue
            
            row_data = {}
            for idx, field_name in field_indices.items():
                value = row[idx] if idx < len(row) else None
                row_data[field_name] = value
            
            # 如果提供了schema，进行数据验证
            if schema:
                try:
                    validated = schema(**row_data)
                    row_data = validated.model_dump()
                except Exception:
                    continue  # 跳过验证失败的行
            
            result.append(row_data)
        
        wb.close()
        return result
    
    @classmethod
    def generate_template(
        cls,
        columns: Dict[str, str],
        sheet_name: str = "Sheet1"
    ) -> BytesIO:
        """
        生成导入模板
        :param columns: 列映射，格式为 {字段名: 显示名}
        :param sheet_name: 工作表名称
        :return: Excel模板文件的BytesIO对象
        """
        return cls.export_to_excel([], columns, sheet_name)


@dataclass
class SummaryData:
    """摘要页数据"""
    basic_info: List[Dict[str, Any]]
    peak_range: List[Dict[str, Any]]
    steady_range: List[Dict[str, Any]]
    metric_unit: str
    metric_columns: List[str]


@dataclass
class DetailData:
    """详细页数据"""
    sheet_name: str
    columns: List[str]
    data: List[List[Any]]
